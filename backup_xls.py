# /usr/bin/env python
# -*- utf-8 -*-

import os
from openpyxl import load_workbook
from datetime import date, timedelta

xls_path = 'E:\\运维管理\\巡检报告\\数据备份记录'
all_day = []
def get_day():
    today = date.today()
    firstday = date(2017, 10, 20)
    delta = timedelta(1)
    day = firstday
    while day <= today:
        if day.weekday() < 5:
            all_day.append(day)
        day = day + delta

def create_xls():
    for date in all_day:
        os.system('E: && cd %s && copy 备份记录2017-07-31.xlsx 备份记录%s.xlsx' % (xls_path, date))

def read_xls():
    for day in all_day:
        #print(day)
        file = '%s\\备份记录%s.xlsx' % (xls_path, day)
        wb = load_workbook(filename=file)
        ws = wb.get_sheet_by_name('Sheet1')  # 获取特定的 worksheet
        # 通过坐标读取值
        ws.cell(row=1, column=4).value = '%s' % day
        ws.cell(row=4, column=4).value = 'YXT-%s.sql' % day
        ws.cell(row=5, column=4).value = 'YXT%s.bak' % day
        ws.cell(row=6, column=4).value = 'YXT_YZG_DATA_CENTER_backup_%s.bak' % day
        ws.cell(row=7, column=4).value = 'YXT_HIS_backup_%s.bak' % day
        print(ws.cell(row=7, column=4).value)
        wb.save(file)

#write_xls()

get_day()
#create_xls()
read_xls()
